import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime

class CadastroApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Cadastro de Clientes e Vendas")
        self.root.geometry("1800x1000")

        # Configurações para a tabela de clientes
        self.tree_clientes = ttk.Treeview(root, columns=("ID", "Nome", "Telefone"), show="headings")
        self.tree_clientes.place(x=900, y=250, width=800, height=600)
        self.tree_clientes.heading("ID", text="ID")
        self.tree_clientes.heading("Nome", text="Nome")
        self.tree_clientes.heading("Telefone", text="Telefone")

        # Campos para cadastrar cliente
        self.cadastro_cliente_frame = tk.Frame(root)
        self.cadastro_cliente_frame.place(x=1000, y=10, width=300, height=200)

        self.nome_label = tk.Label(self.cadastro_cliente_frame, text="Nome:")
        self.nome_label.grid(row=0, column=0, padx=30, pady=10)
        self.nome_entry = tk.Entry(self.cadastro_cliente_frame)
        self.nome_entry.grid(row=0, column=1, padx=10, pady=10)

        self.telefone_label = tk.Label(self.cadastro_cliente_frame, text="Telefone:")
        self.telefone_label.grid(row=1, column=0, padx=10, pady=10)
        self.telefone_entry = tk.Entry(self.cadastro_cliente_frame)
        self.telefone_entry.grid(row=1, column=1, padx=10, pady=10)

        self.cadastrar_cliente_button = tk.Button(self.cadastro_cliente_frame, text="Cadastrar Cliente", command=self.salvar_cliente)
        self.cadastrar_cliente_button.grid(row=2, column=0, columnspan=2, pady=10)

        # Configurações para a tabela de vendas
        self.tree_vendas = ttk.Treeview(root, columns=("ID", "Data", "Valor", "Produto"), show="headings")
        self.tree_vendas.place(x=40, y=250, width=800, height=600)
        self.tree_vendas.heading("ID", text="ID")
        self.tree_vendas.heading("Data", text="Data")
        self.tree_vendas.heading("Valor", text="Valor")
        self.tree_vendas.heading("Produto", text="Produto")

        # Campos para cadastrar venda
        self.cadastro_venda_frame = tk.Frame(root)
        self.cadastro_venda_frame.place(x=200, y=10, width=300, height=200)

        self.data_label = tk.Label(self.cadastro_venda_frame, text="Data:")
        self.data_label.grid(row=0, column=0, padx=10, pady=10)
        self.data_entry = tk.Entry(self.cadastro_venda_frame)
        self.data_entry.grid(row=0, column=1, padx=10, pady=10)

        # Preenche o campo de data com a data atual
        self.data_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))

        self.valor_label = tk.Label(self.cadastro_venda_frame, text="Valor:")
        self.valor_label.grid(row=1, column=0, padx=10, pady=10)
        self.valor_entry = tk.Entry(self.cadastro_venda_frame)
        self.valor_entry.grid(row=1, column=1, padx=10, pady=10)

        self.produto_label = tk.Label(self.cadastro_venda_frame, text="Produto:")
        self.produto_label.grid(row=2, column=0, padx=10, pady=10)
        self.produto_entry = tk.Entry(self.cadastro_venda_frame)
        self.produto_entry.grid(row=2, column=1, padx=10, pady=10)

        self.cadastrar_venda_button = tk.Button(self.cadastro_venda_frame, text="Cadastrar Venda", command=self.salvar_venda)
        self.cadastrar_venda_button.grid(row=3, column=0, columnspan=2, pady=10)
        self.listar_clientes()
        self.listar_vendas()

    def salvar_cliente(self):
        nome = self.nome_entry.get()
        telefone = self.telefone_entry.get()

        workbook = self.abrir_planilha("clientes.xlsx")
        sheet = workbook.active

        # Obtém o próximo ID disponível
        next_id = self.get_next_id(sheet)

        sheet.append([next_id, nome, telefone])

        workbook.save("clientes.xlsx")
        self.listar_clientes()

    def salvar_venda(self):
        data = self.data_entry.get()
        valor = self.valor_entry.get()
        produto = self.produto_entry.get()

        workbook = self.abrir_planilha("vendas.xlsx")
        sheet = workbook.active

        # Obtém o próximo ID disponível
        next_id = self.get_next_id(sheet)

        sheet.append([next_id, data, valor, produto])

        workbook.save("vendas.xlsx")
        self.listar_vendas()

    def listar_clientes(self):
        workbook = self.abrir_planilha("clientes.xlsx")
        sheet = workbook.active

        # Limpa a tabela
        self.limpar_tabela(self.tree_clientes)

        rows = sheet.iter_rows(min_row=2, values_only=True)
        for row in sorted(rows, key=lambda x: (int(x[0]) if x[0] else 0, x[1]), reverse=True):
            self.tree_clientes.insert("", "end", values=row)

    def listar_vendas(self):
        workbook = self.abrir_planilha("vendas.xlsx")
        sheet = workbook.active

        # Limpa a tabela
        self.limpar_tabela(self.tree_vendas)

        # Preenche a tabela com os dados de vendas
        rows = sheet.iter_rows(min_row=2, values_only=True)
        for row in sorted(rows, key=lambda x: (int(x[0]) if x[0] else 0, x[1]), reverse=True):
            self.tree_vendas.insert("", "end", values=row)

    def limpar_tabela(self, tree):
        for item in tree.get_children():
            tree.delete(item)

    def abrir_planilha(self, filename):
        try:
            workbook = load_workbook(filename)
        except FileNotFoundError:
            workbook = Workbook()
            workbook.save(filename)
            workbook = load_workbook(filename)
        return workbook

    def get_next_id(self, sheet):
    # Obtém o ID da última linha ou retorna 0 se a planilha estiver vazia
        last_id = sheet.cell(row=sheet.max_row, column=1).value
        return (last_id + 1) if last_id else 1


if __name__ == "__main__":
    root = tk.Tk()
    app = CadastroApp(root)
    root.mainloop()
